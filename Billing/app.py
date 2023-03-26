from flask import Flask, make_response, request, jsonify, send_from_directory
import sqlalchemy, datetime, os, json, requests
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, Table, Column, Integer, String, MetaData, select
import os
import datetime

app = Flask(__name__)
app.config['JSON_SORT_KEYS'] = False

engine = sqlalchemy.create_engine("mysql+pymysql://billdbuser:billdbpass@mysql-server:3306/billdb")

def is_provider_exist(name):
     conn=engine.connect()
     is_exist=conn.execute(sqlalchemy.text(f"select name from Provider where name='{name}'"))
     conn.close()
     if is_exist.first() != None:
          return True
     else:
          return False

def is_provider_id_exist(id):
    conn=engine.connect()
    is_exist=conn.execute(sqlalchemy.text(f"select id from Provider where id={id}"))
    conn.close()
    if is_exist.first() != None:
        return True 
    else:
        return False

def is_truck_id_exist(id):
    conn=engine.connect()
    is_exist=conn.execute(sqlalchemy.text(f"select id from Trucks where id='{id}'"))
    conn.close()
    if is_exist.first() != None:
        return True
    else:
        return False


@app.route('/provider', methods=["POST"])
def post_provider():
    if request.method != "POST":
        return make_response("Method not allowed",405)
    if not request.is_json:
        return make_response("Bad Request",400)
    data=request.json
    if not isinstance(data,dict):
        return make_response("Bad Request",400)
    provider_name=data.get('name')
    if provider_name==None or provider_name=="" or len(provider_name)>255:
        return make_response("Bad Request",400)
    if is_provider_exist(provider_name):
        return make_response("Provider exists", 400)
    conn=engine.connect()
    conn.execute(sqlalchemy.text(f"INSERT INTO Provider (name) VALUES ('{provider_name}')"))
    conn.commit()
    getid=conn.execute(sqlalchemy.text(f"select id from Provider where name='{provider_name}'"))
    conn.close()
    response={"id" : getid.first()[0]}
    return make_response(jsonify(response), 200)


@app.route('/provider/<id>', methods=["PUT"])
def update_provider_name(id):
    if request.method!="PUT":
        return make_response("Method not allowed",405)
    if not request.is_json:
         return make_response("Bad Request",400)
    data=request.json
    if not isinstance(data,dict):
        return make_response("Bad Request",400)
    name_to_update = data.get('name')
    if name_to_update==None or name_to_update=="" or len(name_to_update)>255:
        return make_response("Bad Request",400)
    if not is_provider_id_exist(id):
        return make_response("id does not exist",404)
    conn=engine.connect()
    conn.execute(sqlalchemy.text(f"UPDATE Provider SET name='{name_to_update}' WHERE id={id}"))
    conn.commit()
    conn.close()
    return make_response("Provider id updated", 200)
    

@app.route('/truck', methods=["POST"])
def post_truck():
    if request.method == "POST":
        if not request.is_json:
            return make_response("Bad Request",400)
        data=request.json
        if not isinstance(data,dict):
            return make_response("Bad Request",400)
        provider_id = data.get('provider')
        truck_id = data.get('id')
        if provider_id==None or truck_id==None or provider_id=="" or truck_id=="" or  len(truck_id)>10:
            return make_response("Missing provider or truck ID", 400)
        if not is_provider_id_exist(provider_id) :
            return make_response("Provider not found", 404)
        if is_truck_id_exist(truck_id):
            return make_response("Truck already registered", 400)
        conn=engine.connect()
        conn.execute(sqlalchemy.text(f"INSERT INTO Trucks (id, provider_id) VALUES ('{truck_id}', {provider_id})"))
        conn.commit()
        conn.close()
        return make_response("Truck added successfully", 200)
    else:
        return make_response("Method not allowed",405)


@app.route('/truck/<id>', methods=["GET"])
def truck_tara_and_sessions(id):
    #id is the truck license.
    if not is_truck_id_exist(id):
        return make_response("The truck id is not exist", 404)
    
    t1=request.args.get('from')
    t2=request.args.get('to')

    if t1 == "" or t1==None: 
        #defult time for the bill
        t1=datetime.datetime.now().strftime("%Y%m01000000")
    if t2 == "" or t2==None:
        t2=datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    else:
        #check the right format
        try: 
            time_1 = datetime.datetime.strptime(t1, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            time_2 = datetime.datetime.strptime(t2, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            if time_1 > time_2:
                return make_response("Inadmissible times", 400)

        except ValueError:
            return make_response("Data format unaccepable", 400)
    time_1 = datetime.datetime.strptime(t1, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
    time_2 = datetime.datetime.strptime(t2, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
    try:
        response = requests.get(f'http://3.76.109.165:8083/item/{id}?from={time_1}&to={time_2}')
        json_data = response.json()
        return make_response(jsonify(json_data), 200)
    except:
        return make_response(jsonify("{ }"), response.status_code)


@app.route('/truck/<id>', methods=["PUT"])
def put_truck(id):
    if request.method != "PUT":
        return make_response("Method not allowed", 405)
    if not is_truck_id_exist(id):
        return make_response("truck doesn't exist", 404)
    if not request.is_json:
        return make_response("Bad Request: Content is not json", 400)
    data=request.json
    if not isinstance(data,dict):
        return make_response("Bad Request",400)
    provider_id=data.get('provider')
    if provider_id==None or provider_id=="":
         return make_response('Bad request',400)
    if not is_provider_id_exist(provider_id):
        return make_response("Specified provider doesn't exist", 400)
    conn=engine.connect()
    conn.execute(sqlalchemy.text(f"UPDATE Trucks SET provider_id={provider_id} WHERE id={id}"))
    conn.commit()
    conn.close()
    return make_response("Updated truck provider",200)


@app.route('/rates', methods=["GET","POST"])
def rates():
    # define database metadata
    metadata = MetaData()

    # define Rates table schema
    rates_table = Table('Rates', metadata,
        Column('product_id', String),
        Column('rate', Integer),
        Column('scope', String)
        )
    
    if request.method == "GET":
        # load all records from Rates table
        with engine.connect() as conn:
            select_stmt = select(rates_table)
            result = conn.execute(select_stmt)
            data = result.fetchall()

        # create a new Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active

        # write the data to the worksheet
        ws.append(['product_id', 'rate', 'scope'])
        for row in data:
            ws.append(list(row))

        # save the workbook to the "out" folder
        folder_path = "out"
        filename = os.path.join(folder_path, "export_rates.xlsx")
        wb.save(filename=filename)
        return send_from_directory(os.path.join(app.root_path,"out"),"export_rates.xlsx", as_attachment=True)

    elif request.method == "POST":
        # load Excel file from the "in" folder
        if not request.is_json:
            return make_response("Bad Request: Content isn't json", 400)
        data=request.json
        if not isinstance(data,dict):
            return make_response("Bad Request",400)
        file=data.get('file')
        folder_path = "in"
        if file==None or file == "":
             return make_response("Bad Request", 400)
        filename = os.path.join(folder_path, file)
        if not os.path.isfile(filename):
            return make_response("Bad Request: specified file doesn't exist", 400)
        try:
            wb = load_workbook(filename=filename, read_only=True)
        except:
            return make_response("Bad Request: File type is not supported", 400)
        
        # delete all records from Rates table
        with engine.connect() as conn:
            conn.execute(rates_table.delete())
        
        # insert records from Excel file to Rates table
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            product_id = row[0].value
            rate = row[1].value
            scope = row[2].value
            if product_id is not None and rate is not None and scope is not None:
                if len(str(product_id)) > 50 or len(str(rate)) > 11 or len(str(scope)) > 50:
                    return make_response("Bad request: file values exceeded the allowed length", 400)
                with engine.connect() as conn:
                    conn.execute(rates_table.insert().values(product_id=product_id, rate=rate, scope=scope))
        return make_response("saved", 200)
    else:
        return make_response("Method is not allowed", 405)
            

def js_name(provider_id):
    conn=engine.connect()
    result=conn.execute(sqlalchemy.text(f"SELECT name FROM Provider WHERE id = {provider_id}")).fetchone()
    conn.close()
    prov_name=result[0]  
    return prov_name


def js_truckCount(provider_id):
    conn=engine.connect()

    #count the trucks
    count_result = conn.execute(sqlalchemy.text(f"SELECT COUNT(*) FROM Trucks WHERE provider_id = {provider_id}")).fetchone()
    truck_count = count_result[0]

    #check the ids
    ids_result = conn.execute(sqlalchemy.text(f"SELECT id FROM Trucks WHERE provider_id = {provider_id}")).fetchall()
    truck_ids = set()
    for id in ids_result:
        truck_ids.add(id[0])
    conn.close()
    return truck_count,truck_ids

def js_truck_session_count(truck_ids,t1,t2):
    totalsessions=0
    for id in truck_ids:
        try:
            request=requests.get(f"http://3.76.109.165:8083/item/{id}?from={t1}&to={t2}")
            sessionlist=request.json()["sessions"]
            totalsessions+=len(sessionlist)
        except:
            continue
    return totalsessions

def js_prod_sess(product_id,truck_ids,t1,t2):
    sumkg=0
    sessioncount=0
    truckdict={}
    for id in truck_ids:
        try:
            request=requests.get(f"http://3.76.109.165:8083/item/{id}?from={t1}&to={t2}")
            json_response = request.json()
            if json_response["sessions"] is not None:
                truckdict[id] = set(json_response["sessions"])
        except :
            continue
    try:
        request=requests.get(f"http://3.76.109.165:8083/weight?from={t1}&to={t2}&filter=in")
        json_response = request.json()
        for item in json_response:
            if product_id in item["produce"]:
                for key in truckdict:
                    if item["id"] in truckdict[key] and item["neto"]!="na":
                        sessioncount+=1
                        sumkg+=int(item["neto"])
    except:
        pass
    return sessioncount,sumkg


def js_prod_and_pay(provider_id,truck_ids,t1,t2):
    #set total_pay and products_list
    total_pay = 0

    products_list = []

    conn=engine.connect()

    #making dictionaries of products
    
    result_exact_id=conn.execute(sqlalchemy.text(f"SELECT product_id, scope FROM Rates WHERE scope='{provider_id}'")).fetchall()
    result_all=conn.execute(sqlalchemy.text(f"SELECT product_id, scope FROM Rates WHERE scope='All'")).fetchall()
    #making sure that if a product_id has scope=All and Scope=provider_id
    #the rate that will be selected is the one with Scope=provider_id
    prod_id_scope_dict=dict()
    for row in result_exact_id:
        prod_id_scope_dict[row[0]] = row[1]
    for row in result_all:
        if row[0] not in prod_id_scope_dict:
            prod_id_scope_dict[row[0]] = row[1]

    for key in prod_id_scope_dict:
        prod_sess, amount_kg = js_prod_sess(key,truck_ids,t1,t2)
        rate_res=conn.execute(sqlalchemy.text(f"SELECT rate FROM Rates WHERE product_id='{key}' AND scope='{prod_id_scope_dict[key]}'")).fetchone()
        rate = rate_res[0]
        pay = amount_kg * rate
        total_pay += pay
        if int(prod_sess) > 0:
            products_list.append({"product": key, "count": prod_sess, "amount": amount_kg, "rate": rate, "pay": pay})
    
    conn.close()

    return products_list, total_pay


@app.route('/bill/<id>', methods=["GET"])
def get_bill(id):
    prov_id=id
    if not is_provider_id_exist(prov_id):
        return make_response("provider was not found", 404)

    t1=request.args.get('from')
    t2=request.args.get('to')

    if t1 == "" or t1==None: 
        #defult time for the bill
        t1=datetime.datetime.now().strftime("%Y%m01000000")
    if t2 == "" or t2==None:
        t2=datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    else:
        #check the right format
        try: 
            time_1 = datetime.datetime.strptime(t1, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            time_2 = datetime.datetime.strptime(t2, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")

        except ValueError:
            return make_response("Data format unaccepable", 400)
        if time_1 > time_2:
            return make_response("Inadmissible times", 400)

    #vars
    prov_name = js_name(prov_id)
    time_1 = datetime.datetime.strptime(t1, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
    time_2 = datetime.datetime.strptime(t2, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
    truck_count, truck_ids = js_truckCount(prov_id)
    truck_sessions_count=js_truck_session_count(truck_ids,t1,t2)
    products, total_pay = js_prod_and_pay(prov_id,truck_ids,t1,t2)

    # Create the dictionary
    bill = {
        "id": prov_id, 
        "name": prov_name, 
        "from": time_1, 
        "to": time_2, 
        "truckCount": truck_count, 
        "sessionCount": truck_sessions_count, 
        "products": products,
        "total": total_pay 
    }
    if not bill["products"] and truck_sessions_count == 0:
        bill["error"]="error could not reach weight correctly, partial information only"
    
    bill_json = jsonify(bill)
    return make_response(bill_json,200)



@app.route('/health', methods=["GET"])
def check_health():
        try:
              conn=engine.connect()
              conn.execute(sqlalchemy.text("select 1"))
              conn.close()
              return make_response("OK", 200)
        except:
              return make_response("OK, connection to the Databaes failed", 503)


if __name__ == "__main__":
    app.run(debug=True)

